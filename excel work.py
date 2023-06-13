import openpyxl

def copy_data(source_file, source_sheet, target_file, target_sheet, source_range, target_start_cell):
    # Load the source workbook and sheet
    source_workbook = openpyxl.load_workbook(source_file)
    source_worksheet = source_workbook[source_sheet]

    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    target_worksheet = target_workbook[target_sheet]

    # Split the source range into start and end cells
    source_start_cell, source_end_cell = source_range.split(':')

    # Get the row and column indices of the source range
    start_row, start_column = openpyxl.utils.cell.coordinate_to_tuple(source_start_cell)
    end_row, end_column = openpyxl.utils.cell.coordinate_to_tuple(source_end_cell)

    # Get the row and column indices of the target start cell
    target_start_row, target_start_column = openpyxl.utils.cell.coordinate_to_tuple(target_start_cell)

    # Copy the data from the source range to the target range
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            source_cell = source_worksheet.cell(row=row, column=col)
            target_cell = target_worksheet.cell(
                row=row - start_row + target_start_row,
                column=col - start_column + target_start_column
            )
            target_cell.value = source_cell.value

    # Save the target workbook
    target_workbook.save(target_file)

# Example usage
copy_data('Advanced Management Programme Module 1 ( 17th April, 2023) Feedback Form (Responses).xlsx', 'Form Responses 1', 'test excel.xlsx', 'Input work sheet', 'B2:H17', 'B4')
copy_data('Advanced Management Programme Module 1 ( 17th April, 2023) Feedback Form (Responses).xlsx', 'Form Responses 1', 'test excel.xlsx', 'Input work sheet', 'J2:P17', 'I4')
copy_data('Advanced Management Programme Module 1 ( 17th April, 2023) Feedback Form (Responses).xlsx', 'Form Responses 1', 'test excel.xlsx', '1', 'Q2:Q17', 'C2')
